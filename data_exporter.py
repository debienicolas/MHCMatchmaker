from io import BytesIO, StringIO
import base64
import os
from openpyxl import Workbook
from openpyxl.styles import Font
import csv
import json
import logging

# project imports
from database import get_database

# setting up 
db = get_database()
logger = logging.getLogger(__name__)


def excel_to_base64_csv(workbook):
    """
    Convert and openpyxl workbook to a base64 encoded csv string
    """

    base64_csvs = {}

    # Iterate through all sheets
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        
        # Create a string buffer to hold CSV data
        csv_buffer = StringIO()
        csv_writer = csv.writer(csv_buffer)
        
        # Write sheet data to CSV buffer
        for row in sheet.iter_rows(values_only=True):
            csv_writer.writerow(row)
        
        # Get the CSV string from the buffer
        csv_string = csv_buffer.getvalue()
        
        # Encode the CSV string to base64
        base64_csv = base64.b64encode(csv_string.encode('utf-8')).decode('utf-8')
        
        # Add to the dictionary
        base64_csvs[sheet_name] = base64_csv
    
    return base64_csvs
        

def export_input(donors, recipients, relevant_classes, write_to_file=False):
    """
    Exports the input data to Excel, CSV, and JSON formats.

    This function processes donor and recipient data for each relevant class,
    organizing it into Excel worksheets, CSV rows, and a JSON structure.
    It also adjusts column widths in the Excel file for better readability.

    Args:
        donors (dict): A dictionary containing donor information.
        recipients (dict): A dictionary containing recipient information.
        relevant_classes (list): A list of relevant MHC classes to process.
        write_to_file (bool, optional): If True, writes the data to files in the 'results/export' directory.
                                        Defaults to False.

    Returns:
        result (dict): {"excel": base64 encoded string, "csv": csv data, "json": json data}

    """

    input_wb = Workbook()
    json_data = {}
    csv_data = []

    for cls in relevant_classes:
        input_ws = input_wb.create_sheet(cls)
        input_ws.append(["Donor ID", "Recipient ID", "Haplotype"])
        json_data[cls] = {"donors": {}, "recipients": {}}
        
        for donor in donors:
            haplotype = donors[donor]["classified"][cls]
            row = [donor, "Donor"] + haplotype
            input_ws.append(row)
            csv_data.append([cls] + row)
            json_data[cls]["donors"][donor] = haplotype
            
        for recipient in recipients:
            haplotype = recipients[recipient]["classified"][cls]
            row = [recipient, "Recipient"] + haplotype
            input_ws.append(row)
            csv_data.append([cls] + row)
            json_data[cls]["recipients"][recipient] = haplotype

    # give the columns enough width so the text inside is visible
    for sheet in input_wb.sheetnames:
        input_ws = input_wb[sheet]
        for column in input_ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            input_ws.column_dimensions[column_letter].width = adjusted_width

    # delete the default sheet
    input_wb.remove(input_wb["Sheet"])

    if write_to_file:
        os.makedirs("results/export", exist_ok=True)
        input_wb.save("results/export/input.xlsx")
        
        # Export to CSV
        with open("results/export/input.csv", "w", newline="") as csvfile:
            csv_writer = csv.writer(csvfile)
            csv_writer.writerow(["Class", "ID", "Type", "Haplotype"])
            csv_writer.writerows(csv_data)
        
        # Export to JSON
        with open("results/export/input.json", "w") as jsonfile:
            json.dump(json_data, jsonfile, indent=2)

    # convert the excel file to a base64 encoded string
    input_buffer = BytesIO()
    input_wb.save(input_buffer)
    input_buffer.seek(0)
    input_base64 = base64.b64encode(input_buffer.getvalue()).decode()

    return {
        "excel": input_base64,
        "csv": excel_to_base64_csv(input_wb)
    }


def export_alignment(donors, recipients, relevant_classes, write_to_file=False):
    """
    Exports the alignment data to Excel, CSV, and JSON formats.

    This function processes alignment data for each relevant class, organizing it into
    Excel worksheets, CSV rows, and a JSON structure. It includes a consensus sequence
    for each class and aligned sequences for all donor and recipient alleles.

    Args:
        donors (dict): A dictionary containing donor information.
        recipients (dict): A dictionary containing recipient information.
        relevant_classes (list): A list of relevant MHC classes to process.
        write_to_file (bool, optional): If True, writes the data to files in the 'results/export' directory.
                                        Defaults to False.

    Returns:
        dict: {"excel": base64 encoded string, "csv": csv data, "json": json data}
    """

    alignment_wb = Workbook()

    for cls in relevant_classes:
        alignment_ws = alignment_wb.create_sheet(cls)
        consensus_seq = db.get_consensus_seq(cls)

        header = ["Allele ID"]
        for i in range(len(consensus_seq)):
            header.append(i+1)
        alignment_ws.append(header)

        # add the consensus sequence to the header
        consensus = ["Consensus"] + list(consensus_seq)
        alignment_ws.append(consensus)

        # make it so that the consensus sequence and the header row stay at the top when scrolling
        alignment_ws.freeze_panes = "A3"

        # make consensus sequence bold
        for cell in alignment_ws[1]:
            cell.font = Font(bold=True)
        for cell in alignment_ws[2]:
            cell.font = Font(bold=True)

        
        for donor in donors:
            for allele in donors[donor]["classified"][cls]:
                # Fetch allele data from MongoDB
                allele_data = db.find(allele)
                aligned_seq = allele_data.aligned_seq
                row = [allele] + list(aligned_seq)
                alignment_ws.append(row)
    
        for recipient in recipients:
            for allele in recipients[recipient]["classified"][cls]:
                # Fetch allele data from MongoDB
                allele_data = db.find(allele)
                aligned_seq = allele_data.aligned_seq
                row = [allele] + list(aligned_seq)
                alignment_ws.append(row)
    
    # give the first column enough width so the text inside is visible and make the rest of the columns narrower (3 symbols wide)
    for sheet in alignment_wb.sheetnames:
        alignment_ws = alignment_wb[sheet]
        max_length = 0
        for cell in alignment_ws['A']:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        alignment_ws.column_dimensions['A'].width = max_length + 2  # Add some padding

        for column in alignment_ws.columns:
            column_letter = column[0].column_letter
            if column_letter != "A":
                alignment_ws.column_dimensions[column_letter].width = 5
    
    # delete the default sheet
    alignment_wb.remove(alignment_wb["Sheet"])
    

    # Turn the alignment data into a csv
    csv_data = []
    for sheet in alignment_wb.sheetnames:
        alignment_ws = alignment_wb[sheet]
        for row in alignment_ws.iter_rows(values_only=True):
            csv_data.append(row)

    if write_to_file:
        os.makedirs("results/export", exist_ok=True)
        alignment_wb.save("results/export/alignment.xlsx")

        with open("results/export/alignment.csv", "w", newline="") as csvfile:
            csv_writer = csv.writer(csvfile)
            csv_writer.writerows(csv_data)

       

    # convert the excel file to a base64 encoded string, for easy communication to frontend
    alignment_buffer = BytesIO()
    alignment_wb.save(alignment_buffer)
    alignment_buffer.seek(0)
    alignment_base64 = base64.b64encode(alignment_buffer.getvalue()).decode()

    return {
        "excel": alignment_base64,
        "csv": excel_to_base64_csv(alignment_wb),
    }


def export_sas_scores(sas_scores, relevant_classes, write_to_file=False):
    """
    Exports the individual allele SAS scores and the grouped SAS scores to Excel, CSV.

    Args:
        sas_scores (dict): A dictionary containing SAS scores for each allele and class.
        relevant_classes (list): A list of relevant MHC classes to process.
        write_to_file (bool, optional): If True, writes the data to files in the 'results/export' directory.
                                        Defaults to False.

    Returns:
        dict: {"excel": base64 encoded string, "csv": csv data}
    """

    sas_scores_wb = Workbook()

    random_donor_id = list(sas_scores.keys())[0]
    
    for cls in relevant_classes:
        ws = sas_scores_wb.create_sheet(cls)
        
        # Get all unique positions
        all_positions = range(len(sas_scores[random_donor_id][cls]))

        # Create header row
        header = ["Donor/Recipient ID"] + [f"{pos+1}" for pos in all_positions]
        ws.append(header)

        # Add data for each donor
        for id in sas_scores:
            row = [id]
            for pos in all_positions:
                if sas_scores[id][cls] != {}:
                    row.append(sas_scores[id][cls][pos]["rsa"])
                else:
                    break
            ws.append(row)

        # Adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width

    # Remove the default sheet
    sas_scores_wb.remove(sas_scores_wb["Sheet"])

   
    if write_to_file:
        os.makedirs("results/export", exist_ok=True)
        sas_scores_wb.save("results/export/sas_scores.xlsx")
        
    # Convert the Excel file to a base64 encoded string
    sas_scores_buffer = BytesIO()
    sas_scores_wb.save(sas_scores_buffer)
    sas_scores_buffer.seek(0)
    sas_scores_base64 = base64.b64encode(sas_scores_buffer.getvalue()).decode()

    return {
        "excel": sas_scores_base64,
        "csv": excel_to_base64_csv(sas_scores_wb)
    }
    

def export_mismatches(difference_scores, relevant_classes, write_to_file=False,output_folder="results/export"):
    """
    This method exports the difference_scores to an excel file.

    :param difference_scores: dict of difference scores
    :param relevant_classes: list of relevant classes

    :return: return a base64 encoded string of the excel file
    """

    recipients = list(difference_scores.keys())
    donors = list(difference_scores[recipients[0]].keys())


    mismatches_wb = Workbook()

    for cls in relevant_classes:
        mismatches_ws = mismatches_wb.create_sheet(cls)
        
        length = len(difference_scores[recipients[0]][donors[0]][cls]["updated_mismatches"])

        # Add header row with recipient-donor pair labels
        header = ['Recipient-Donor'] + [i+1 for i in range(length)]
        mismatches_ws.append(header)

        for recip in recipients:
            for donor in donors:
                row = [f"{recip}-{donor}: initial donor mismatches"]
                mismatches = difference_scores[recip][donor][cls]["donor_diff"]
                for i in mismatches:
                    if i is list():
                        row.append(None)
                    else:
                        row.append(",".join(i))

                mismatches_ws.append(row)

                row = [f"{recip}-{donor}: SAS filtered donor mismatches"]
                mismatches = difference_scores[recip][donor][cls]["updated_mismatches"]
                for i in mismatches:
                    if i is list():
                        row.append(None)
                    else:
                        row.append(",".join(i))
                mismatches_ws.append(row)
            
                row = [f"{recip}-{donor}: initial recipient mismatches"]
                mismatches = difference_scores[recip][donor][cls]["recip_diff"]
                for i in mismatches:
                    if i is list():
                        row.append(None)
                    else:
                        row.append(",".join(i))
                mismatches_ws.append(row)
                
                row = [f"{recip}-{donor}: SAS filtered recipient mismatches"]
                mismatches = difference_scores[recip][donor][cls]["updated_recip_mismatches"]
                for i in mismatches:
                    if i is list():
                        row.append(None)
                    else:
                        row.append(",".join(i))
                mismatches_ws.append(row)

                # add an empty row
                mismatches_ws.append([])
                
    # give the first column enough width so the text inside is visible and make the rest of the columns narrower (3 symbols wide)
    for sheet in mismatches_wb.sheetnames:
        mismatches_ws = mismatches_wb[sheet]
        max_length = 0
        for cell in mismatches_ws['A']:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        mismatches_ws.column_dimensions['A'].width = max_length + 2  # Add some padding
    

    mismatches_wb.remove(mismatches_wb["Sheet"])
    if write_to_file:
        os.makedirs(output_folder, exist_ok=True)
        mismatches_wb.save(os.path.join(output_folder, "mismatches.xlsx"))


    # Convert the Excel file to a base64 encoded string
    mismatches_buffer = BytesIO()
    mismatches_wb.save(mismatches_buffer)
    mismatches_buffer.seek(0)
    mismatches_base64 = base64.b64encode(mismatches_buffer.getvalue()).decode()


    return {
        "excel": mismatches_base64,
        "csv": excel_to_base64_csv(mismatches_wb)
    }
                

def export_known_eplets(known_eplets, relevant_classes, write_to_file=False,output_path="results/export"):
    """
    This method exports the known eplets to a json file.

    :param known_eplets: dict of known eplets
    :param relevant_classes: list of relevant classes
    :param write_to_file: bool, if True, write the data to a json file
    :param output_path: str, path to the output file
    """

    eplets_wb = Workbook()

    for cls in relevant_classes:
        ws = eplets_wb.create_sheet(cls)

        # Add header row with recipient-donor pair labels
        header = ["Recipient-Donor", "Known Eplets"]
        ws.append(header)

        for recip in known_eplets:
            for donor in known_eplets[recip]:
                # Donor diff eplets
                row = [f"{recip}-{donor} donor mismatches"]
                if cls in known_eplets[recip][donor].keys():
                    eplets = list(known_eplets[recip][donor][cls]["donor_diff"].keys())
                    for eplet in eplets:
                        row.append(eplet)
                ws.append(row)

                # Recip diff eplets
                row = [f"{recip}-{donor} recipient mismatches"]
                if cls in known_eplets[recip][donor].keys():
                    eplets = list(known_eplets[recip][donor][cls]["recip_diff"].keys())
                    for eplet in eplets:
                        row.append(eplet)
                ws.append(row)
    

    for sheet in eplets_wb.sheetnames:
        eplets_ws = eplets_wb[sheet]
        max_length = 0
        for cell in eplets_ws['A']:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        eplets_ws.column_dimensions['A'].width = max_length + 2  # Add some padding

    eplets_wb.remove(eplets_wb["Sheet"])
    if write_to_file:
        os.makedirs(output_path, exist_ok=True)
        eplets_wb.save(os.path.join(output_path, "eplets.xlsx"))


    # Convert the Excel file to a base64 encoded string
    eplets_buffer = BytesIO()
    eplets_wb.save(eplets_buffer)
    eplets_buffer.seek(0)
    eplets_base64 = base64.b64encode(eplets_buffer.getvalue()).decode()


    return {
        "excel": eplets_base64,
        "csv": excel_to_base64_csv(eplets_wb)
    }
            


def create_entity_info(donors,recipients,write_to_file=False,output_path="results/") -> dict:
    """
    Collects relevant information for each allele and stores it in a dictionary.

    Parameters:
        donors (dict): A dictionary containing donor information.
        recipients (dict): A dictionary containing recipient information.

    Returns:
        entity_info (dict): A dictionary containing the entity information.
    """

    entity_info = {}

    donors_and_recips = {**donors, **recipients}
    for id in donors_and_recips:
        for allele in donors_and_recips[id]["Haplotype"]:
            if allele not in entity_info:
                # Fetch allele data from MongoDB
                allele_data = db.find(allele)
                
                rsa_scores = allele_data.aligned_rsa

                # Turn the rsa_scores into a dictionary with the keys as positions and the values as the rsa scores
                rsa_scores_dict = {i+1: rsa_scores[i] for i in range(len(rsa_scores))}
                
                entity_info[allele] = {
                    "linked_id": [id],
                    "seq": allele_data.sequence,
                    "seq_length": len(allele_data.sequence),
                    "aligned_seq": allele_data.aligned_seq,
                    "aligned_seq_length": len(allele_data.aligned_seq),
                    "rsa_scores": rsa_scores_dict
                }
            else:
                entity_info[allele]["linked_id"].append(id)
    
    if write_to_file:
        os.makedirs(output_path, exist_ok=True)
        with open(output_path + "entity_info.json", "w") as f:
            json.dump(entity_info, f)


    return entity_info


def get_raw_mismatches(donors,recipients,difference_scoring):
        """
        Unused method
        """
        # This method should be used after the initial MHC matchmaker algorithm (before Solven accessibility filtering)
        # mismatches are either donor_diffs or recip_diffs

        raw_mismatches = {}
        
        for recipient_id in recipients.keys():
            raw_mismatches[recipient_id] = {}
            for donor_id in donors.keys():
                # check if self.difference_scoring has been calculated for donor_id and recipient_id
                if donor_id not in difference_scoring[recipient_id]:
                    logging.info("Difference scoring not calculated for donor_id: %s and recipient_id: %s", donor_id, recipient_id)
                    logging.info("Calculating difference scoring for donor_id: %s and recipient_id: %s", donor_id, recipient_id)
                    raise ValueError("Difference scoring not calculated for donor_id: %s and recipient_id: %s", donor_id, recipient_id)
                
                donor_diff = difference_scoring[recipient_id][donor_id]["donor_diff"]
                recip_diff = difference_scoring[recipient_id][donor_id]["recip_diff"]


                mismatches = {}

                for i, (d, r) in enumerate(zip(donor_diff, recip_diff)):
                    # donor_diff and recip_diff are lists of lists
                    # check for donor_diff
                    mis_str = []
                    if d != list():
                        mis_str += d
                    if r != list():
                        mis_str += r
                    
                    if mis_str != list():
                        mismatches[i] = mis_str

                raw_mismatches[recipient_id][donor_id] = mismatches
        return raw_mismatches
    

def generate_ranking_data(donors, recipients, difference_scoring):
    """

    
    """
    ranking_data = {}

    random_donor_id = list(donors.keys())[0]

    for recip_id in recipients.keys():
        recipient_data = {"recipientID": recip_id, "scores": {}}

        for clas in difference_scoring[recip_id][random_donor_id]:
            recipient_data["scores"][clas] = []
    
            for donor_id in donors.keys():
    
                if donor_id not in difference_scoring[recip_id]:
                    raise ValueError(f"Difference scoring not calculated for donor_id: {donor_id} and recipient_id: {recip_id}")
                
                donor_diff_score = difference_scoring[recip_id][donor_id][clas]["donor_diff_score"]
                donor_diff_total = len(recipients[recip_id]["haplotypeClassGrouped"][clas])
                
                if donor_diff_total != 0:
                    percent = round(donor_diff_score / donor_diff_total, 4) * 100.0 
                    percent = 100 - percent
                    updated_score = difference_scoring[recip_id][donor_id][clas]["updated_mismatches_count"]
                    updated_score_percent = 100 - round(updated_score / donor_diff_total, 4) * 100.0
                else:
                    percent = None
                    updated_score = None
                    updated_score_percent = None
                
                total_sequence_length = len(difference_scoring[recip_id][donor_id][clas]["updated_mismatches"])


                recipient_data["scores"][clas].append({
                    "donorID": donor_id, 
                    "score": percent,
                    "mismatches_donor": len([mismatch for mismatch in difference_scoring[recip_id][donor_id][clas]["donor_diff"] if mismatch != list()]),
                    "mismatches_recip": len([mismatch for mismatch in difference_scoring[recip_id][donor_id][clas]["recip_diff"] if mismatch != list()]),
                    "updated_mismatches_donor": len([mismatch for mismatch in difference_scoring[recip_id][donor_id][clas]["updated_mismatches"] if mismatch != list()]),
                    "updated_mismatches_recip": len([mismatch for mismatch in difference_scoring[recip_id][donor_id][clas]["updated_recip_mismatches"] if mismatch != list()]),
                    "total_sequence_length": total_sequence_length ,
                    "updated_score": updated_score_percent,
                    "updated_mismatches": difference_scoring[recip_id][donor_id][clas]["updated_mismatches"]})
            
        ranking_data[recip_id] = recipient_data
    
    return ranking_data


def generate_ranking_data_csv(ranking_data):
    """
    Generate a csv file from the ranking data for every recipient
    """
    pass





def get_aligned_seqs(donors, recipients):
    """
    Get the aligned sequences for all alleles in the donors and recipients.
    """
    aligned_seqs = {}
    donors_and_recips = {**donors, **recipients}
    
    for entity in donors_and_recips:
        for allele in donors_and_recips[entity]["Haplotype"]:
            # Fetch allele data from MongoDB
            allele_data = db.find(allele)
            
            aligned_seq = allele_data.aligned_seq
            if aligned_seq is not None:
                aligned_seqs[allele] = aligned_seq
            else:
                logger.warning(f"Aligned sequence not found for allele {allele}")
    
    return aligned_seqs
    
    
def generate_output_excel_files(donors,recipients,difference_scoring,sas_scores,relevant_classes,known_eplets):
        """
        Generates output excel files:
        - input.xlsx: contains the input data
        - alignment.xlsx: contains the aligned sequences

        Needs relevant classes, donors, recipients, difference

        """

        return {
            "input": export_input(donors, recipients, relevant_classes, write_to_file=True),
            "alignment": export_alignment(donors, recipients, relevant_classes, write_to_file=True),
            "sas_scores": export_sas_scores(sas_scores, relevant_classes, write_to_file=True),
            "mismatches": export_mismatches(difference_scoring, relevant_classes, write_to_file=True),
            "eplets": export_known_eplets(known_eplets, relevant_classes, write_to_file=True),
        }


if __name__ == "__main__":

    from matchmaker import MHCMatchmaker

    mm = MHCMatchmaker(output_path="example_results")
    mm.perform_matching(input_filename="test_data/Worked_out_example.xlsx")

    generate_output_excel_files(mm.donors,mm.recipients,mm.difference_scoring,mm.sas_scores,mm.get_relevant_classes(),mm.known_eplets)