from typing import Dict, List, Tuple
import logging
# from dotenv import load_dotenv
import json
import os
import pandas as pd
from ast import literal_eval
import openpyxl
import time
from operator import add

# project imports
from database import get_database
from utils.utils import parse_allele_name
from utils.epletMatching import create_eplet_dict

# set up logging
logging.basicConfig(level=logging.INFO, 
                    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
                    datefmt='%Y-%m-%d %H:%M:%S')

logger = logging.getLogger(__name__)


class MHCMatchmaker:
    """
    This is the main class for the MHCMatchmaker.

    Attributes:
        donors (Dict): Dictionary mapping donor IDs to their haplotype information.
        recipients (Dict): Dictionary mapping recipient IDs to their haplotype information.
        difference_scoring (Dict): Dictionary storing the mismatch scores between donors and recipients.
        sas_scores (Dict): Dictionary storing solvent accessibility scores for donors and recipients.
        output_path (str): Path to the output directory. Defaults to "results/".
        db: Database connection object.
        local_db (Dict): Local cache of allele data from the database.
        invalid_alleles (List): List of alleles that were found to be invalid.
        transformed_alleles (Dict): Dictionary mapping original allele names to transformed names.
        known_eplets (Dict): Dictionary storing information about known eplets found in the analysis.
    """

    def __init__(self, output_path:str = "results/"):
        self.donors = {}
        self.recipients = {}
        self.difference_scoring = {}
        self.sas_scores = {}

        self.output_path = output_path
        # make the output directory if it does not exist
        os.makedirs(self.output_path, exist_ok=True)

        # Setup the database, either local of mongoD
        self.db = get_database()

        # put all the relevant information from the database in a local dictionary
        self.local_db = {}
    
    def load_local_db(self):
        """
        Loads the relevant info for this session from the database into a local dictionary.
        
        This method retrieves allele data for all alleles in both donor and recipient
        haplotypes and stores them in the local_db dictionary for faster access.
        """
        # recipients
        for recip_id in self.recipients:
            for allele in self.recipients[recip_id]["Haplotype"]:
                allele_data = self.db.find(allele)
                self.local_db[allele] = allele_data
        # donors
        for donor_id in self.donors:
            for allele in self.donors[donor_id]["Haplotype"]:
                allele_data = self.db.find(allele)
                self.local_db[allele] = allele_data

        logger.info(f"Local database loaded with {len(self.local_db)} alleles")
        return None

    def set_inputs_csv(self,input_df: pd.DataFrame) -> Tuple[Dict, Dict]:
        """ 
        Loads donor and recipient information from a pandas DataFrame.
        
        The input DataFrame should have the following columns:
        - identifier (str): Unique identifier for the donor or recipient
        - type (str): Either "Donor" or "Recipient"
        - haplotype (List[str]): List of allele IDs that form the haplotype

        Parameters: 
            input_df (pd.DataFrame): DataFrame containing donor and recipient information
        
        Returns:
            Tuple[Dict, Dict]: A tuple containing (donors, recipients) dictionaries
        
        Raises:
            ValueError: If input data is empty or all haplotypes are empty
            AssertionError: If input data is in the wrong format or missing required columns
        """

        # set empty donors and recipients for this class
        self.donors = {}
        self.recipients = {}


        # check if the input is empty, no columns or no rows
        if input_df.empty:
            raise ValueError("Input is empty")
        
        # Make sure the input is acceptable and in the right format
        assert "identifier" in input_df.columns, "Input must contain an identifier column"
        assert "type" in input_df.columns, "Input must contain a type column"
        assert "haplotype" in input_df.columns, "Input must contain a haplotype column"

        # make sure there are no extra columns
        assert input_df.columns.tolist() == ["identifier", "type", "haplotype"], "Input must contain only the identifier, type and haplotype columns"

        # make sure all donor ids are unique
        assert input_df.identifier.duplicated().sum() == 0, "Donor names must be unique"

        # make sure all recipient ids are unique
        assert input_df.identifier.duplicated().sum() == 0, "Recipient names must be unique"

        # check if there are donors and recipients in the input file
        assert "Donor" in input_df.type.unique(), "Input must contain donor type"
        assert "Recipient" in input_df.type.unique(), "Input must contain recipient type"
        # make sure there are no other values in the type column than "Donor" or "Recipient"
        assert input_df.type.isin(["Donor", "Recipient"]).all(), "Input must contain only the Donor and Recipient types"

        # check if there is at least one donor and one recipient
        assert len(input_df[input_df.type == "Donor"]) > 0, "Input must contain at least one donor"
        assert len(input_df[input_df.type == "Recipient"]) > 0, "Input must contain at least one recipient"

        # create a dictionary of donors and recipients
        for _, row in input_df.iterrows():
            if row['type'] == 'Donor':
                self.donors[row["identifier"]] = {"Haplotype": row["haplotype"]}
            elif row['type'] == 'Recipient':
                self.recipients[row["identifier"]] = {"Haplotype": row["haplotype"]}
        
        # raise an exception if all the donor and recipient haplotypes are empty
        all_donor_empty = all(self.donors[donor]["Haplotype"] == [] for donor in self.donors)
        all_recipient_empty = all(self.recipients[recipient]["Haplotype"] == [] for recipient in self.recipients)
        if all_donor_empty and all_recipient_empty:
            raise ValueError("All donor and recipient haplotypes are empty")

        # log the amount of donors and recipients
        logger.info(f"{len(self.donors)} donors loaded")
        logger.info(f"{len(self.recipients)} recipients loaded")

        
        return self.donors, self.recipients

    def set_inputs_excel(self,workbook: openpyxl.Workbook) -> Tuple[Dict, Dict]:
        """
        Loads donor and recipient information from an Excel workbook.
        
        The Excel file should have the following format:
        - First column: identifier (str) - Unique identifier for the donor or recipient
        - Second Column: type (str) - Either "Donor" or "Recipient"
        - Following columns: Allele IDs that form the haplotype

        Parameters: 
            workbook (openpyxl.Workbook): Excel workbook object containing donor and recipient information
        
        Returns:
            Tuple[Dict, Dict]: A tuple containing (donors, recipients) dictionaries
        
        Raises:
            AssertionError: If no donors or recipients are found, or if identifiers are duplicated
        """

        # reset the donors and recipients for this class
        self.donors = {}
        self.recipients = {}

        sheet = workbook.active

        # set the max row and column
        max_row = sheet.max_row
        max_col = sheet.max_column

        donors = {}
        recipients = {}
        # loop over the rows and add the identifier, type and haplotype to the donors and recipients dictionaries
        for i in range(2,max_row+1):
            identifier = sheet.cell(row=i, column=1).value
            type = sheet.cell(row=i, column=2).value

            # check that if the identifier is not empty it must be Donor or Recipient
            if type is not None:
                assert type in ["Donor", "Recipient"], "The type must be Donor or Recipient"
            
            # Donor and recipient names must be unique
            if type == "Donor":
                assert identifier not in donors, f"Donor {identifier} is duplicated"
            elif type == "Recipient":
                assert identifier not in recipients, f"Recipient {identifier} is duplicated"
            

            # for the haplotype column, build a list with all the values in the remaining columns that are not empty
            haplotype = []
            for j in range(3,max_col+1):
                if sheet.cell(row=i, column=j).value is not None:
                    # strip the value of any whitespace
                    haplotype.append(sheet.cell(row=i, column=j).value.strip())
            
            if type == "Donor":
                donors[identifier] = {"Haplotype": haplotype}
            elif type == "Recipient":
                recipients[identifier] = {"Haplotype": haplotype}
        
        assert len(donors) > 0, "No donors found in the input data"
        assert len(recipients) > 0, "No recipients found in the input data"

        # Donor names must be unique
        
        # log the amount of donors and recipients
        logger.info(f"{len(donors)} donors loaded")
        logger.info(f"{len(recipients)} recipients loaded")

        self.donors = donors
        self.recipients = recipients

        return (donors, recipients)

    def check_alleles(self) -> None:
        """
        Validates all alleles from the input and attempts to correct invalid ones.
        
        This method performs several checks on each allele:
        1. Verifies if the allele exists in the database
            2. If not found, checks if it's a secondary name of another allele
            3. If still not found, attempts to find a similar base allele
        4. Verifies the allele has necessary attributes
        
        Invalid alleles are removed from haplotypes and stored in self.invalid_alleles.
        Transformed alleles are tracked in self.transformed_alleles.
        
        Raises:
            ValueError: If critical allele validation fails
        """
        # list to keep track of possible invalid alleles
        self.invalid_alleles = []
        # list ot keep track of the transformed alleles
        self.transformed_alleles = {}

        donors_and_recips = {**self.donors, **self.recipients}

        for id in donors_and_recips:
            # copy the haplotype to avoid modifying the original
            haplotype = donors_and_recips[id]["Haplotype"].copy()
            for allele in haplotype:
                valid_allele = True

                # Check if the allele is in the database
                allele_data = self.db.find(allele)

                if allele_data is None:

                    # Check if the allele is in the secondary names of any other allele
                    #secondary_name_match = self.collection.find_one({"secondary_names": allele})
                    secondary_name_match = self.db.specific_find("secondary_names", allele)

                    if secondary_name_match:
                        # Replace with the known name
                        known_name = secondary_name_match["_id"]
                        if id in self.donors:
                            # replace the allele in the haplotype
                            self.donors[id]["Haplotype"].remove(allele)
                            self.donors[id]["Haplotype"].append(known_name)
                            logger.info(f"Allele {allele} not found in database, but {known_name} found")
                        elif id in self.recipients:
                            self.recipients[id]["Haplotype"].remove(allele)
                            self.recipients[id]["Haplotype"].append(known_name)
                            logger.info(f"Allele {allele} not found in database, but {known_name} found")
                        allele_data = secondary_name_match
                        allele_data = self.db.bson_to_dataclass(secondary_name_match)

                        self.transformed_alleles[allele] = known_name

                    else:
                        # e.g. A*01:01:01:01 -> only the A*01:01 is really important
                        # If the allele is not found in the database, check if the allele can be changed with a similar base allele
                        base_allele = parse_allele_name(allele)[:3]
                        base_allele = base_allele[0] + "*" + ":".join(base_allele[1:])
                        # find all alleles that have base_allele in their name (_id)
                        #similar_alleles = self.collection.find({"$text": {"$search": base_allele}})
                        similar_alleles = self.db.str_in_allele(base_allele)
                        similar_alleles = list(similar_alleles)

                        # if there is only one similar allele, use it
                        if len(similar_alleles) == 1 and similar_alleles[0]["status"] != "abandoned" and similar_alleles[0]["aligned_seq"] != "":
                            logger.info(f"Allele {allele} not found in database, but similar allele {similar_alleles[0]['_id']} found")
                            new_allele_id = similar_alleles[0]["_id"]
                            if id in self.donors:
                                # replace the allele in the haplotype
                                self.donors[id]["Haplotype"].remove(allele)
                                self.donors[id]["Haplotype"].append(new_allele_id)
                            elif id in self.recipients:
                                self.recipients[id]["Haplotype"].remove(allele)
                                self.recipients[id]["Haplotype"].append(new_allele_id)
                            allele_data = self.db.bson_to_dataclass(similar_alleles[0])
                            self.transformed_alleles[allele] = new_allele_id

                        # if there are multiple similar alleles,use the shortest one
                        elif len(similar_alleles) > 1:
                            logger.info(f"Allele {allele} not found in database, but multiple similar alleles found: {[allele['_id'] for allele in similar_alleles]}")
                            # find the shortest allele
                            sorted_similar_alleles = sorted(similar_alleles, key=lambda x: len(x["_id"]))
                            #print("Sorted similar alleles: ", sorted_similar_alleles)
                            # choose the first one if the status is not "abandoned" and the sequence is not empty
                            for similar_allele in sorted_similar_alleles:
                                if similar_allele["status"] != "abandoned" and similar_allele["aligned_seq"] != "":
                                    shortest_allele = similar_allele
                                    break
                            new_allele_id = shortest_allele["_id"]
                            if id in self.donors:
                                self.donors[id]["Haplotype"].remove(allele)
                                self.donors[id]["Haplotype"].append(new_allele_id)
                            elif id in self.recipients:
                                self.recipients[id]["Haplotype"].remove(allele)
                                self.recipients[id]["Haplotype"].append(new_allele_id)
                            allele_data = self.db.bson_to_dataclass(shortest_allele)
                            self.transformed_alleles[allele] = new_allele_id
                        else:
                            valid_allele = False
                            logger.warning(f"Allele {allele} not found in database and no similar allele found")
                            #raise ValueError(f"Allele {allele} not found in database and no similar allele found")

                #assert isinstance(allele_data, Allele), f"{allele} data is not of type Allele, it is {type(allele_data)}"
                if valid_allele:
                    # Check if the allele has all the necessary attributes
                    if allele_data.aligned_seq == "" or not allele_data.aligned_seq:
                        #logger.warning(f"Allele {allele} has no aligned sequence")
                        pass
                        #raise ValueError(f"Allele {allele} has no aligned sequence")
                    if allele_data.aligned_rsa is None:
                        logger.warning(f"Allele {allele} has no aligned rsa")
                        #raise ValueError(f"Allele {allele} has no aligned rsa")
                    
                if not valid_allele:
                    # remove the allele from the haplotype
                    if id in self.donors:
                        self.donors[id]["Haplotype"].remove(allele)
                    elif id in self.recipients:
                        self.recipients[id]["Haplotype"].remove(allele)

                    self.invalid_alleles.append(allele)
        
        self.load_local_db()
                    
        #print("Invalid alleles: ", self.invalid_alleles)
        return
    
    def classify_haplotypes(self) -> None:
        """
        Classifies alleles in haplotypes into specific HLA classes.
        
        This method categorizes each allele in donor and recipient haplotypes into:
        - Class I
        - Class II DQA
        - Class II DQB
        - Class II DRA
        - Class II DRB
        
        The classification is stored in a new "classified" key in the donors and recipients dictionaries:
        e.g., self.donors[donor_id]["classified"] = {"I": [...], "IIDQA": [...], ...}
        
        Returns:
            Tuple[Dict, Dict]: Updated donors and recipients dictionaries with classified haplotypes
        
        Raises:
            ValueError: If an allele cannot be classified or is not found in the database
        """

        donors_recipients = {**self.donors, **self.recipients}
        for id in donors_recipients:
            classes = {"I":[], "IIDQA":[], "IIDQB":[], "IIDRA":[], "IIDRB":[]}
            for allele in donors_recipients[id]["Haplotype"]:
                # Fetch allele data from MongoDB
                #allele_data = self.collection.find_one({"_id": allele})
                allele_data = self.local_db[allele]
                
                if allele_data is None:
                    logger.warning(f"Allele {allele} not found in database")
                    raise ValueError(f"Allele {allele} not found in database")
                
                allele_class = allele_data.allele_class
                locus = allele_data.locus
                
                if allele_class == "I":
                    classes["I"].append(allele)
                elif allele_class == "II":
                    if "DQA" in locus:
                        classes["IIDQA"].append(allele)
                    elif "DQB" in locus:
                        classes["IIDQB"].append(allele)
                    elif "DRA" in locus:
                        classes["IIDRA"].append(allele)
                    elif "DRB" in locus:
                        classes["IIDRB"].append(allele)
                    else:
                        logger.warning(f"{allele} could not be classified")
                        raise ValueError(f"{allele} could not be classified into one of these classes: {list(classes.keys())}")
                else:
                    logger.warning(f"{allele} could not be classified")
                    raise ValueError(f"{allele} could not be classified into one of these classes: {list(classes.keys())}")
                
            # do a sanity check: the length of the haplotype should be equal to the sum of all classes list
            assert len(donors_recipients[id]["Haplotype"]) == sum([len(classes[key]) for key in classes.keys()])
            # add the classes to the donors and recipients dictionary
            if id in self.donors:
                self.donors[id]["classified"] = classes
            elif id in self.recipients:
                self.recipients[id]["classified"] = classes
        
        logger.info("Haplotypes have been classified")
        
        return self.donors, self.recipients

    def group_alleles(self) -> None:
        """
        Groups the alleles of each haplotype by class for donors and recipients.
        
        This method creates a combined sequence representation for each class of alleles.
        For each class with multiple alleles, it combines their sequences by position.
        
        The grouped sequences are stored in a new "haplotypeClassGrouped" key in the 
        donors and recipients dictionaries.
        
        Returns:
            Tuple[Dict, Dict]: Updated donors and recipients dictionaries with grouped haplotypes
        """

        donors_and_recips = {**self.donors, **self.recipients}
        # loop over all donors and recipients
        for id in donors_and_recips:
            haplotypeClassGrouped = {}
            # loop over all classes
            for allele_class in donors_and_recips[id]["classified"]:
                haplotype = donors_and_recips[id]["classified"][allele_class]   
                # if the haplotype is empty, add an empty list
                if len(haplotype) == 0:
                    haplotypeClassGrouped[allele_class] = []
                elif len(haplotype) == 1:
                    #allele_data = self.collection.find_one({"_id": haplotype[0]})
                    allele_data = self.local_db[haplotype[0]]
                    base_sequence = allele_data.aligned_seq
                    haplotypeClassGrouped[allele_class] = list(base_sequence)
                else:
                    #allele_data = self.collection.find_one({"_id": haplotype[0]})
                    allele_data = self.local_db[haplotype[0]]
                    base_sequence = allele_data.aligned_seq
                    for allele in haplotype[1:]:
                        #allele_data = self.collection.find_one({"_id": allele})
                        allele_data = self.local_db[allele]
                        base_sequence = list(map(add, base_sequence, allele_data.aligned_seq))
                    haplotypeClassGrouped[allele_class] = base_sequence
            
            if id in self.donors:
                self.donors[id]["haplotypeClassGrouped"] = haplotypeClassGrouped
            elif id in self.recipients:
                self.recipients[id]["haplotypeClassGrouped"] = haplotypeClassGrouped
        
        # logging
        logger.info("Alleles have been grouped")
        
        return self.donors, self.recipients
                
    def calcSingleDifference(self, donor_id: str, recipient_id: str) -> Dict:
        """
        Calculates the mismatches between a specific donor and recipient.
        
        This method compares the grouped haplotypes of a donor and recipient and identifies
        differences at each position. It calculates both donor-to-recipient and 
        recipient-to-donor differences.
        
        Parameters:
            donor_id (str): The donor identifier
            recipient_id (str): The recipient identifier
        
        Returns:
            Dict: Dictionary containing difference scores and details for each class:
                - donor_diff: Elements in donor not in recipient
                - recip_diff: Elements in recipient not in donor
                - donor_diff_score: Number of positions with donor differences
                - recip_diff_score: Number of positions with recipient differences
                - all_donor_diff_counts: Counts of each differing element in donor
                - all_donor_diff_ratios: Ratios of each differing element in donor
                - all_recip_diff_counts: Counts of each differing element in recipient
                - all_recip_diff_ratios: Ratios of each differing element in recipient
        
        Raises:
            AssertionError: If donor or recipient is not found, or if they have different classes
        """

        # sanity checks
        assert donor_id in self.donors, f"Donor {donor_id} not found in donors"
        assert recipient_id in self.recipients, f"Recipient {recipient_id} not found in recipients"
        assert self.donors[donor_id]["classified"].keys() == self.recipients[recipient_id]["classified"].keys(), "Donor and recipient have different classes"

        # Loop over all classes and calculate the difference scores
        class_scores = {}
        for clas in self.recipients[recipient_id]["classified"].keys():
            donor_grouped = self.donors[donor_id]["haplotypeClassGrouped"][clas]
            recipient_grouped = self.recipients[recipient_id]["haplotypeClassGrouped"][clas]
            
            # in the case that the donor and recipient have different lengths, we need to pad the shorter one with '-'
            donor_grouped = [allele if i < len(donor_grouped) else '-' for i, allele in enumerate(donor_grouped)]
            recipient_grouped = [allele if i < len(recipient_grouped) else '-' for i, allele in enumerate(recipient_grouped)]

            # calculate the difference between the donor and recipient grouped haplotypes
            all_donor_diff_counts = []
            all_donor_diff_ratios = []
            all_recip_diff_counts = []
            all_recip_diff_ratios = []

            for a1, a2 in zip(donor_grouped, recipient_grouped):
                donor_diff_elems = set(a1).difference(set(a2)) # elements in a1 that are not in a2

                donor_diff_counts = {elem: a1.count(elem) for elem in donor_diff_elems} 
                donor_diff_ratios = {elem: a1.count(elem) / len(a1) for elem in donor_diff_elems}
                
                all_donor_diff_counts.append(donor_diff_counts)
                all_donor_diff_ratios.append(donor_diff_ratios)

                recip_diff_elems = set(a2).difference(set(a1))
                recip_diff_counts = {elem: a2.count(elem) for elem in recip_diff_elems}
                recip_diff_ratios = {elem: a2.count(elem) / len(a2) for elem in recip_diff_elems}

                all_recip_diff_counts.append(recip_diff_counts)
                all_recip_diff_ratios.append(recip_diff_ratios)


            donor_diff = [set(a1).difference(set(a2)) for a1, a2 in zip(donor_grouped, recipient_grouped)] # set(donor_grouped) - set(recipient_grouped)  elements present in donor_grouped but not in recipient_grouped
            recip_diff = [set(a2).difference(set(a1)) for a1, a2 in zip(donor_grouped, recipient_grouped)] # set(recipient_grouped) - set(donor_grouped)  elements present in recipient_grouped but not in donor_grouped
        
            # use a mapping function to map the list to different values:
            # if the set is empty, map to []
            # if the set is not empty, map to the list of the set
            donor_diff = list(map(lambda x: [] if x == set() else list(x), donor_diff))
            recip_diff = list(map(lambda x: [] if x == set() else list(x), recip_diff))

            # calculate the difference scores: if ther diff list is not empty, add 1 to the score
            donor_diff_score = sum([1 for diff in donor_diff if diff != list()])
            recip_diff_score = sum([1 for diff in recip_diff if diff != list()])

            class_scores[clas] = {"donor_diff": donor_diff,
                                  "recip_diff": recip_diff,
                                  "donor_diff_score": donor_diff_score,
                                  "recip_diff_score": recip_diff_score,
                                  "all_donor_diff_counts": all_donor_diff_counts,
                                  "all_donor_diff_ratios": all_donor_diff_ratios,
                                  "all_recip_diff_counts": all_recip_diff_counts,
                                  "all_recip_diff_ratios": all_recip_diff_ratios}


        return class_scores
    
    def calcMHCDifference(self) -> Dict:
        """
        Calculates difference scores for all donor-recipient pairs.
        
        This method iterates through all possible donor-recipient combinations and
        calculates their mismatch scores using calcSingleDifference. The results
        are stored in the self.difference_scoring dictionary.
        
        Returns:
            Dict: Dictionary mapping recipient IDs to dictionaries of donor IDs and their
                 corresponding difference scores
        """

        # Loop over all possible recipient and donor pairs
        for recip_id in self.recipients.keys():
            scores_dict = {}
            for donor_id in self.donors.keys():
                class_scores = self.calcSingleDifference(donor_id, recip_id)
                scores_dict[donor_id] = class_scores

                logger.info("Difference scoring calculated for donor_id: %s and recipient_id: %s", donor_id, recip_id)
            self.difference_scoring[recip_id] = scores_dict
        
        return self.difference_scoring

    def average_sas_scores(self) -> Dict:
        """
        Calculates average solvent accessibility scores for each position in grouped alleles.
        
        This method computes the average relative solvent accessibility (RSA) and 
        absolute solvent accessibility (ASA) scores for each position in the grouped
        alleles of donors and recipients. Positions with None values (gaps) are excluded
        from the average calculation.
        
        Returns:
            Dict: Dictionary mapping donor/recipient IDs to dictionaries of classes and
                 position-specific average scores:
                 {donor_id/recipient_id: {class: {position: {rsa: float, asa: float, total: int}}}}
        
        Raises:
            ValueError: If an allele is not found or has no aligned RSA/ASA values
        """

        donors_and_recips = {**self.donors, **self.recipients}
        all_sas_scores = {}
        # loop over all donors and recipients
        for id in donors_and_recips:
            sas_scores = {} 
            # loop over all the classes
            for clas in donors_and_recips[id]["classified"]:
                class_scores = {}
                # loop over all the allels in the classified haplotype
                for allele in donors_and_recips[id]["classified"][clas]:

                    # Fetch allele data 
                    #allele_data = self.collection.find_one({"_id": allele})
                    allele_data = self.local_db[allele]
                    
                    if allele_data is None:
                        logger.warning(f"Allele {allele} not found in database")
                        raise ValueError(f"Allele {allele} not found in database")
                    
                    rsa_scores = allele_data.aligned_rsa
                    asa_scores = allele_data.aligned_asa
                    
                    if rsa_scores is None:
                        logger.warning(f"Allele {allele} has no aligned rsa")
                        raise ValueError(f"Allele {allele} has no aligned rsa")
                    if asa_scores is None:
                        logger.warning(f"Allele {allele} has no aligned asa")
                        raise ValueError(f"Allele {allele} has no aligned asa")

                    
                    for i, (r, a) in enumerate(zip(rsa_scores, asa_scores)):
                        if i not in class_scores:
                            class_scores[i] = {"rsa": 0, "asa": 0, "total": 0}
                        if r is not None and a is not None:
                            class_scores[i]["rsa"] += r
                            class_scores[i]["asa"] += a
                            class_scores[i]["total"] += 1
                
                if class_scores:
                    class_scores = {int(i): {
                                                "rsa": valuedict["rsa"] / valuedict["total"] if valuedict["total"] != 0 else None,
                                                "asa": valuedict["asa"] / valuedict["total"] if valuedict["total"] != 0 else None,
                                                "total": valuedict["total"]
                                                } for i, valuedict in class_scores.items()}
                sas_scores[clas] = class_scores
            all_sas_scores[id] = sas_scores
        
        # write the results to a json file
        with open(os.path.join(self.output_path, "sas_scores.json"), "w") as f:
            json.dump(all_sas_scores, f)
        
        self.sas_scores = all_sas_scores

        # Logging
        logger.info("SAS scores calculated and saved to %s", self.output_path + "sas_scores.json")

        return self.sas_scores
    
    def filter_by_sas(self, rsa_threshold: float = 0.5) -> Dict:
        """
        Filters mismatches based on solvent accessibility scores.
        
        This method examines the initial mismatches between donor-recipient pairs and
        filters out those where the relative solvent accessibility (RSA) score is below
        the specified threshold. Mismatches with low RSA scores are considered less
        immunologically relevant.
        
        Parameters:
            rsa_threshold (float): The RSA threshold below which mismatches are filtered out.
                                   Defaults to 0.5.
        
        Returns:
            Dict: Updated difference_scoring dictionary with filtered mismatches and counts:
                 {recipient_id: {donor_id: {class: {
                     updated_mismatches: [...],
                     updated_mismatches_count: int,
                     updated_recip_mismatches: [...],
                     updated_recip_mismatches_count: int
                 }}}}
        """
        
        # loop over each recipient and donor pair, and each class for that pair
        for recip in self.recipients:
            for donor in self.donors:
                for clas in self.difference_scoring[recip][donor]:
                    # initial mismatches
                    mismatches = self.difference_scoring[recip][donor][clas]["donor_diff"]
                    updated_mismatches = []
                    
                    for index,value in enumerate(mismatches):
                        if value != list():
                            # check rsa from self.donor_sas_scores and self.recipient_sas_scores
                            rsa = self.sas_scores[donor][clas][index]["rsa"]

                            if rsa is not None and rsa < rsa_threshold:
                                updated_mismatches.append([])
                            else:
                                updated_mismatches.append(value)
                        else:
                            updated_mismatches.append([])
                    # updated mismatches is a list of lists: either an empty list (no mismatches) or a list of mismatches
                    self.difference_scoring[recip][donor][clas]["updated_mismatches"] = updated_mismatches
                    self.difference_scoring[recip][donor][clas]["updated_mismatches_count"] = sum([1 for diff in updated_mismatches if diff != list()])

                    recip_mismatches = self.difference_scoring[recip][donor][clas]["recip_diff"]
                    updated_recip_mismatches = []
                    for index,value in enumerate(recip_mismatches):
                        if value != list():
                            rsa = self.sas_scores[recip][clas][index]["rsa"]
                            if rsa is not None and rsa < rsa_threshold:
                                updated_recip_mismatches.append([])
                            else:
                                updated_recip_mismatches.append(value)
                        else:
                            updated_recip_mismatches.append([])
                    self.difference_scoring[recip][donor][clas]["updated_recip_mismatches"] = updated_recip_mismatches
                    self.difference_scoring[recip][donor][clas]["updated_recip_mismatches_count"] = sum([1 for diff in updated_recip_mismatches if diff != list()])
                    

                logger.info("Solvent accessibile filtered mismatches for recipient %s and donor %s ", recip, donor)

        # write the results to a json file
        with open(os.path.join(self.output_path, "difference_scores.json"), "w") as f:
            json.dump(self.difference_scoring, f)

        logger.info("Solvent accessibile filtered mismatches for all recipient and donor pairs")
        logger.info("Filtered mismatches saved to %s", self.output_path + "difference_scores.json")

        return self.difference_scoring

    def has_eplet(self, seq: str, eplet_data: Dict) -> bool:
        """
        Determines if a sequence contains a specific eplet.
        
        An eplet is present if all specified amino acids are found at their
        respective positions in the sequence.
        
        Parameters:
            seq (str): The amino acid sequence to check
            eplet_data (Dict): Dictionary mapping positions to expected amino acids
                              {position(str): amino_acid(str)}
        
        Returns:
            bool: True if the sequence contains the eplet, False otherwise
        """
        eplet_positions = sorted(eplet_data.keys())
        for pos in eplet_positions:
            if int(pos) >= len(seq):
                return False
            if seq[int(pos)-1] != eplet_data[pos]:
                return False
        return True

    def check_known_eplets(self) -> Dict:
        """
        Identifies known HLA eplets in the mismatches between donors and recipients.
        
        This method checks if mismatch positions correspond to known eplet positions.
        If a match is found, it verifies if the complete eplet is present in the donor
        or recipient alleles. The analysis is performed for both donor-to-recipient and
        recipient-to-donor differences.
        
        Returns:
            Dict: Dictionary mapping recipient IDs to dictionaries of donor IDs and
                 the eplets found in their mismatches:
                 {recipient_id: {donor_id: {class: {
                     donor_diff: {eplet_id: {...}},
                     recip_diff: {eplet_id: {...}}
                 }}}}
        """

        file_paths = {
            "IIDRB": "data/eplets/eplets_DRB.json",
            "IIDQ": "data/eplets/eplets_DQ.json",
            "I": "data/eplets/eplets_I.json"
        }


        # This method assumes the eplet positions are the same for the mismatch positions        
        
        random_donor_id = list(self.donors.keys())[0]
        classes = self.donors[random_donor_id]["classified"].keys()

        local_db = {}

        # load the local db with the recipient alleles
        for recip_id in self.recipients.keys():
            for clas in classes:
                for allele in self.recipients[recip_id]["classified"][clas]:
                    allele_data = self.local_db[allele]
                    local_db[allele] = allele_data
        
        # load the local db with the donor alleles
        for donor_id in self.donors.keys():
            for clas in classes:
                for allele in self.donors[donor_id]["classified"][clas]:
                    allele_data = self.local_db[allele]
                    local_db[allele] = allele_data


        eplets_found = {}

        # Loop over all the recipients and donor pairings and check for known eplets
        for recip_id in self.recipients.keys():
            eplets_found[recip_id] = {}
            for donor_id in self.donors.keys():
                
                class_eplets_found = {}
                # loop over all the classes
                for clas in classes:
                    
                    if clas not in file_paths:
                        if clas == "IIDQA" or clas == "IIDQB":
                            file_path = file_paths["IIDQ"]
                        else:
                            continue
                    else:
                        file_path = file_paths[clas]
                    
                    # load the eplet data
                    with open(file_path, "r") as f:
                        eplets_dict = json.load(f)

                    
                    # get all the positions where possible eplets are found
                    eplet_pos = create_eplet_dict(eplets_dict)
                    # eplet_pos is a dictionary with the first eplet position as keys and lists of eplet ids as values
                    

                    # DONOR DIFF
                    mismatches = self.difference_scoring[recip_id][donor_id][clas]["donor_diff"]
                    # turn mismatches into a dictionary {position: mismatch} for mismatches != list()
                    mismatches_dict = {index+1: mismatch for index, mismatch in enumerate(mismatches) if mismatch != list()}
                    # filter by taking out ['-']
                    mismatches_dict = {str(k): v for k, v in mismatches_dict.items() if '-' not in v}
                    donor_known_eplets = {}
                    # loop ove all the mismatches 
                    tried_eplets = set()
                    for position, mismatch in mismatches_dict.items():
                        # if the position of the mismatch is in the eplet_pos dictionary
                        if position in eplet_pos:
                            eplet_ids = eplet_pos[position]
                            for eplet_id in eplet_ids:
                                if eplet_id in tried_eplets:
                                    continue
                                tried_eplets.add(eplet_id)
                                eplet_data = eplets_dict[eplet_id]
                                # check in the allele of the donor of that class to see if any allele has the eplet
                                donor_alleles = self.donors[donor_id]["classified"][clas]
                                donors_that_have_eplet = []
                                for allele in donor_alleles:
                                    # Fetch allele data from MongoDB
                                    #allele_data = self.db.find(allele)
                                    allele_data = local_db[allele]
                                    if allele_data.eplets is not None:
                                        if eplet_id in allele_data.eplets:
                                            donors_that_have_eplet.append(allele)
                                    else:
                                        if self.has_eplet(allele_data.aligned_seq, eplet_data):
                                            donors_that_have_eplet.append(allele)
                                recipients_that_have_diff = []
                                for recip_allele in self.recipients[recip_id]["classified"][clas]:
                                    #recip_allele_data = self.db.find(recip_allele)
                                    recip_allele_data = local_db[recip_allele]
                                    if recip_allele_data.aligned_seq[int(position)-1] != mismatch:
                                        recipients_that_have_diff.append(recip_allele)
                                
                                if any(donors_that_have_eplet) and any(recipients_that_have_diff):
                                    buffer = 2
                                    # get min and max position of the eplet from the eplet_data
                                    eplet_min_pos = min([int(pos) for pos in eplet_data.keys()]) - buffer
                                    eplet_max_pos = max([int(pos) for pos in eplet_data.keys()]) + buffer
                                    
                                    donor_known_eplets[eplet_id] = {
                                        "donors": donors_that_have_eplet,
                                        "recipients": recipients_that_have_diff,
                                        "mismatch_position": position,
                                        "min_pos": eplet_min_pos,
                                        "max_pos": eplet_max_pos,
                                        "eplet_data": eplet_data
                                    }
                                    

                    # RECIPIENT DIFF                    
                    mismatches = self.difference_scoring[recip_id][donor_id][clas]["recip_diff"]
                    # turn mismatches into a dictionary {position: mismatch} for mismatches != list()
                    mismatches_dict = {index+1: mismatch for index, mismatch in enumerate(mismatches) if mismatch != list()}
                    # filter by taking out ['-']
                    mismatches_dict = {str(k): v for k, v in mismatches_dict.items() if '-' not in v}

                    recip_known_eplets = {}
                    # loop ove all the mismatches 
                    tried_eplets = set()
                    for position, mismatch in mismatches_dict.items():
                        # if the position of the mismatch is in the eplet_pos dictionary
                        if position in eplet_pos:
                            eplet_ids = eplet_pos[position]
                            for eplet_id in eplet_ids:
                                if eplet_id in tried_eplets:
                                    continue
                                tried_eplets.add(eplet_id)
                                eplet_data = eplets_dict[eplet_id]
                                # check in the allele of the donor of that class to see if any allele has the eplet
                                recipient_alleles = self.recipients[recip_id]["classified"][clas]
                                recipients_that_have_eplet = []
                                for allele in recipient_alleles:
                                    # Fetch allele data from MongoDB
                                    #allele_data = self.db.find(allele)
                                    allele_data = local_db[allele]
                                    # if self.has_eplet(allele_data.aligned_seq, eplet_data):
                                    #     recipients_that_have_eplet.append(allele)
                                    if allele_data.eplets is not None:
                                        if eplet_id in allele_data.eplets:
                                            recipients_that_have_eplet.append(allele)
                                    else:
                                        if self.has_eplet(allele_data.aligned_seq, eplet_data):
                                            recipients_that_have_eplet.append(allele)
                                donors_that_have_diff = []
                                for donor_allele in self.donors[donor_id]["classified"][clas]:
                                    #donor_allele_data = self.db.find(donor_allele)
                                    donor_allele_data = local_db[donor_allele]
                                    if donor_allele_data.aligned_seq[int(position)-1] != mismatch:
                                        donors_that_have_diff.append(donor_allele)
                                
                                if any(donors_that_have_diff) and any(recipients_that_have_eplet):
                                    buffer = 2
                                    # get min and max position of the eplet from the eplet_data
                                    eplet_min_pos = min([int(pos) for pos in eplet_data.keys()]) - buffer
                                    eplet_max_pos = max([int(pos) for pos in eplet_data.keys()]) + buffer

                                    recip_known_eplets[eplet_id] = {
                                        "donors": donors_that_have_diff,
                                        "recipients": recipients_that_have_eplet,
                                        "mismatch_position": position,
                                        "min_pos": eplet_min_pos,
                                        "max_pos": eplet_max_pos,
                                        "eplet_data": eplet_data
                                    }

                    if donor_known_eplets or recip_known_eplets:
                        logger.info(f"Known eplets for recipient {recip_id} and donor {donor_id} in class {clas})")
                        class_eplets_found[clas] = {
                            "donor_diff": donor_known_eplets,
                            "recip_diff": recip_known_eplets
                        }
                eplets_found[recip_id][donor_id] = class_eplets_found
        
        with open(os.path.join(self.output_path, "eplets_found.json"), "w") as f:
            json.dump(eplets_found, f)

        return eplets_found

    def get_relevant_classes(self) -> List[str]:
        """
        Identifies HLA classes that are relevant for the matching process.
        
        A class is considered relevant if at least one donor or recipient has
        alleles in that class.
        
        Returns:
            List[str]: List of relevant class identifiers
        """
        random_donor_id = list(self.donors.keys())[0]
        classes = list(self.donors[random_donor_id]["classified"].keys())
        relevant_classes = []
        entities = {**self.donors,**self.recipients}
        for clas in classes:
            relevant = False
            for id in entities:
                if entities[id]["classified"][clas] != []:
                    relevant = True
                    break
            if relevant:
                relevant_classes.append(clas)
        return relevant_classes
    
    def perform_matching(self, input_filename: str) -> Dict:
        """
        Performs the complete MHC matching process from start to finish.
        
        This method executes the full matching workflow:
        1. Loads donor and recipient data from the input file
        2. Validates and corrects alleles
        3. Classifies haplotypes by HLA class
        4. Groups alleles within each class
        5. Calculates mismatches between donors and recipients
        6. Computes average solvent accessibility scores
        7. Filters mismatches based on solvent accessibility
        8. Identifies known eplets in the mismatches
        
        Parameters:
            input_filename (str): Path to the input file (CSV or Excel)
        
        Returns:
            Dict: The final difference scoring dictionary with all mismatch information
        
        Raises:
            ValueError: If the input file format is invalid
        """

        start = time.time()

        if input_filename is not None:

            # check the input file if it is a csv or a excel
            if input_filename.endswith('.csv'):
                df = pd.read_csv(input_filename)
                df.haplotype = df.haplotype.apply(literal_eval)
                donors, recipients = self.set_inputs_csv(df)
            elif input_filename.endswith('.xlsx') or input_filename.endswith('.xls'):
                workbook = openpyxl.load_workbook(input_filename)
                donors, recipients = self.set_inputs_excel(workbook)
            else:
                raise ValueError("Invalid input file format. Please provide a CSV or Excel file.")

        # Check the alleles of the donors and recipients
        self.check_alleles()

        # Classify the input haplotypes
        self.classify_haplotypes()

        # Filter the mismatches by SAS
        self.group_alleles()

        # Calculate the difference scores
        self.calcMHCDifference()

        # average the sas scores
        self.average_sas_scores()

        # Filter the donors by SAS scores
        self.filter_by_sas(rsa_threshold=0.25)

        # check for known eplets
        self.known_eplets = self.check_known_eplets()

        stop = time.time()
        execution_time = stop - start
        print(f"Execution time: {execution_time} seconds")

        return self.difference_scoring
    
        

